import os, io, zipfile, csv, tempfile
from typing import List
import pandas as pd
from celery import Celery, current_task

from core import xml_rows_to_dataframe, iter_xml_rows
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook

# Use REDIS_URL env var if present (Render: set in both web & worker services)
REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379/0")

celery = Celery("tasks", broker=REDIS_URL, backend=REDIS_URL)
celery.conf.update(task_track_started=True)

@celery.task(bind=True)
def convert_task(self, file_ids: List[str], file_names: List[str], columns: List[str], output_format: str):
    """
    Convert uploaded XMLs (by saved file_id) into chosen format.
    If multiple files -> create a zip; single file -> write single artifact.
    Returns JSON with 'filename' for friendly download name.
    """
    os.makedirs("outputs", exist_ok=True)
    total = len(file_ids)
    job_id = self.request.id

    if total > 1:
        zip_path = f"outputs/{job_id}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for idx, (fid, fname) in enumerate(zip(file_ids, file_names), start=1):
                xml_path = f"temp_uploads/{fid}.xml"
                if not os.path.exists(xml_path):
                    raise FileNotFoundError(f"Missing uploaded file: {xml_path}")
                base = os.path.splitext(os.path.basename(fname))[0]
                if output_format == "csv":
                    # Stream CSV rows directly into the zip entry to reduce memory
                    buf = io.StringIO()
                    writer = None
                    # Build header from first row, or from provided columns
                    if columns is not None:
                        header = list(columns)
                        if not header:
                            # nothing to write for this file
                            continue
                    else:
                        # peek first row to get keys
                        first_row = None
                        for r in iter_xml_rows(xml_path):
                            first_row = r
                            break
                        if first_row is None or not first_row:
                            header = []
                        else:
                            header = list(first_row.keys())
                        # re-iterate including first row
                        def row_iter():
                            if first_row is not None and any((first_row.get(k) is not None and first_row.get(k) != "") for k in header):
                                yield {k: first_row.get(k) for k in header}
                            for r in iter_xml_rows(xml_path):
                                if any((r.get(k) is not None and r.get(k) != "") for k in header):
                                    yield {k: r.get(k) for k in header}
                        rows_it = row_iter()
                    if columns is not None:
                        def rows_filtered():
                            for r in iter_xml_rows(xml_path):
                                if any((r.get(k) is not None and r.get(k) != "") for k in header):
                                    yield {k: r.get(k) for k in header}
                        rows_it = rows_filtered()
                    writer = csv.DictWriter(buf, fieldnames=header, extrasaction='ignore')
                    if header:
                        writer.writeheader()
                        for r in rows_it:
                            writer.writerow(r)
                        if buf.tell() > 1_000_000:  # flush in chunks ~1MB
                            zf.writestr(f"{base}.csv", buf.getvalue().encode("utf-8"))
                            buf.seek(0); buf.truncate(0)
                    if buf.tell():
                        zf.writestr(f"{base}.csv", buf.getvalue().encode("utf-8"))
                elif output_format == "parquet":
                    # Stream to a temp parquet file, then add to zip
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".parquet")
                    tmp_path = tmp.name
                    tmp.close()
                    # Determine schema
                    header = columns if columns else None
                    # Build writer after first batch to infer schema
                    writer = None
                    batch_size = 20000
                    rows_buffer = []
                    def normalize_row(row, header_keys):
                        if header_keys is None:
                            header_keys = list(row.keys())
                        out = {k: row.get(k) for k in header_keys}
                        return {k: (None if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)) for k, v in out.items()}
                    for r in iter_xml_rows(xml_path):
                        # establish header if not provided
                        if header is None:
                            header = list(r.keys())
                        rows_buffer.append(normalize_row(r, header))
                        if len(rows_buffer) >= batch_size:
                            table = pa.Table.from_pylist(rows_buffer)
                            if writer is None:
                                writer = pq.ParquetWriter(tmp_path, table.schema)
                            writer.write_table(table)
                            rows_buffer.clear()
                    if rows_buffer:
                        table = pa.Table.from_pylist(rows_buffer)
                        if writer is None:
                            writer = pq.ParquetWriter(tmp_path, table.schema)
                        writer.write_table(table)
                        rows_buffer.clear()
                    if writer is not None:
                        writer.close()
                    # Add to zip and remove temp file
                    zf.write(tmp_path, arcname=f"{base}.parquet")
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
                else:  # xlsx
                    # Use openpyxl write-only mode to stream rows into a temp file
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                    tmp_path = tmp.name
                    tmp.close()
                    wb = Workbook(write_only=True)
                    ws = wb.create_sheet("Rows")
                    # determine header
                    header = list(columns) if columns is not None else None
                    first_row = None
                    it = iter_xml_rows(xml_path)
                    for r in it:
                        first_row = r
                        break
                    if header is None:
                        header = list(first_row.keys()) if first_row else []
                    if header:
                        ws.append(header)
                        # write first row and rest
                        if first_row is not None:
                            row_vals = [first_row.get(k) for k in header]
                            if any((v is not None and v != "") for v in row_vals):
                                ws.append(row_vals)
                        for r in iter_xml_rows(xml_path):
                            row_vals = [r.get(k) for k in header]
                            if any((v is not None and v != "") for v in row_vals):
                                ws.append(row_vals)
                    wb.save(tmp_path)
                    wb.close()
                    zf.write(tmp_path, arcname=f"{base}.xlsx")
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass

                # progress
                percent = int((idx / total) * 100)
                current_task.update_state(state="PROGRESS", meta={
                    "current": idx, "total": total, "file": fname, "progress": percent
                })
        return {"filename": "converted_files.zip"}

    # Single file
    fid, fname = file_ids[0], file_names[0]
    xml_path = f"temp_uploads/{fid}.xml"
    with open(xml_path, "rb") as f:
        data = f.read()
    base = os.path.splitext(os.path.basename(fname))[0]
    if output_format == "csv":
        out_path = f"outputs/{job_id}.csv"
        # stream to CSV directly
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = None
            if columns is not None:
                header = list(columns)
                if not header:
                    return {"filename": f"{base}.csv"}
                writer = csv.DictWriter(f, fieldnames=header, extrasaction='ignore')
                writer.writeheader()
                for r in iter_xml_rows(xml_path):
                    if any((r.get(k) is not None and r.get(k) != "") for k in header):
                        writer.writerow({k: r.get(k) for k in header})
            else:
                first_row = None
                for r in iter_xml_rows(xml_path):
                    first_row = r
                    break
                if first_row is None or not first_row:
                    pd.DataFrame([]).to_csv(out_path, index=False)
                else:
                    header = list(first_row.keys())
                    writer = csv.DictWriter(f, fieldnames=header, extrasaction='ignore')
                    writer.writeheader()
                    if any((first_row.get(k) is not None and first_row.get(k) != "") for k in header):
                        writer.writerow({k: first_row.get(k) for k in header})
                    for r in iter_xml_rows(xml_path):
                        if any((r.get(k) is not None and r.get(k) != "") for k in header):
                            writer.writerow({k: r.get(k) for k in header})
        result_name = f"{base}.csv"
    elif output_format == "parquet":
        out_path = f"outputs/{job_id}.parquet"
        header = columns if columns else None
        writer = None
        batch_size = 20000
        rows_buffer = []
        def normalize_row(row, header_keys):
            if header_keys is None:
                header_keys = list(row.keys())
            out = {k: row.get(k) for k in header_keys}
            return {k: (None if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)) for k, v in out.items()}
        for r in iter_xml_rows(xml_path):
            if header is None:
                header = list(r.keys())
            rows_buffer.append(normalize_row(r, header))
            if len(rows_buffer) >= batch_size:
                table = pa.Table.from_pylist(rows_buffer)
                if writer is None:
                    writer = pq.ParquetWriter(out_path, table.schema)
                writer.write_table(table)
                rows_buffer.clear()
        if rows_buffer:
            table = pa.Table.from_pylist(rows_buffer)
            if writer is None:
                writer = pq.ParquetWriter(out_path, table.schema)
            writer.write_table(table)
            rows_buffer.clear()
        if writer is not None:
            writer.close()
        result_name = f"{base}.parquet"
    else:
        out_path = f"outputs/{job_id}.xlsx"
        # openpyxl write-only streaming
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Rows")
        header = list(columns) if columns is not None else None
        first_row = None
        it = iter_xml_rows(xml_path)
        for r in it:
            first_row = r
            break
        if header is None:
            header = list(first_row.keys()) if first_row else []
        if header:
            ws.append(header)
            if first_row is not None:
                row_vals = [first_row.get(k) for k in header]
                if any((v is not None and v != "") for v in row_vals):
                    ws.append(row_vals)
            for r in iter_xml_rows(xml_path):
                row_vals = [r.get(k) for k in header]
                if any((v is not None and v != "") for v in row_vals):
                    ws.append(row_vals)
        wb.save(out_path)
        wb.close()
        result_name = f"{base}.xlsx"
    return {"filename": result_name}